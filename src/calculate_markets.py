import os
import sys
import csv
import datetime
from collections import OrderedDict

from openpyxl import Workbook, load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = "../data/interim/market_history.xlsx"
market_metadata = {
    "Dow Jones Industrial Average": ("Dow", "../data/raw/market_history/^DJI.csv"),
    "S&P 500": ("SP500", "../data/raw/market_history/^GSPC.csv"),
    "Nasdaq Composite": ("Nasdaq", "../data/raw/market_history/^IXIC.csv"),
}

market_data = {}

for index_name, data in market_metadata.items():
    abbrev = data[0]
    filepath = data[1]

    with open(filepath, "r") as market_file:
        data = csv.reader(market_file)
        headers = next(data)
        date_col = headers.index("Date")
        close_col = headers.index("Close")

        market_data[abbrev] = OrderedDict()

        for row in data:
            date = datetime.date.fromisoformat(row[date_col])
            market_data[abbrev][date] = row[close_col]

intervals = ["", "-3M", "-6M", "-1Y", "-2Y", "-4Y"]

market_wb = Workbook()
market_ws = market_wb.active
market_ws.title = "Market History"
market_ws.cell(row=1, column=1, value="Year")
curr_col = 2
column_lookup = {}
for index_name in market_data:
    for i in intervals:
        label = index_name + i
        market_ws.cell(row=1, column=curr_col, value=label)
        column_lookup[label] = curr_col
        curr_col += 1

curr_row = 2
years = [x for x in range(1990, 2020)]
for year in years:
    market_ws.cell(curr_row, 1, year)
    election_date = datetime.date(year, 11, 1)
    for index_name, data in market_data.items():
        current_index = float(data[election_date])
        market_ws.cell(curr_row, column_lookup[index_name], current_index)

        old_dates = {
            "-3M": datetime.date(year ,   8, 1),
            "-6M": datetime.date(year ,   5, 1),
            "-1Y": datetime.date(year-1, 11, 1),
            "-2Y": datetime.date(year-2, 11, 1),
            "-4Y": datetime.date(year-4, 11, 1),
        }
        for prev_label, prev in old_dates.items():
            prev_index = float(data[prev])
            perc_change = (current_index - prev_index) / prev_index
            market_ws.cell(curr_row, column_lookup[index_name+prev_label], perc_change)

    curr_row += 1


market_wb.save(OUTPUT_FILE)
