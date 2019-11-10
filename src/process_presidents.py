import os
import sys
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

os.chdir(os.path.dirname(os.path.abspath(__file__)))
GOV_PATH = "../data/interim/president_data/"
OUTPUT_PATH = "../data/processed/"

candidates = []
county_votes = []
all_parties = OrderedDict()

# here are separate codes for parties that should
#  be the same (just different abbreviations across years)
equivalents = {
    "Write-In": "Write-ins",
    "Write-in": "Write-ins",
    "American Ind.": "American Independent",
    "National": "National Party",
    "We, the People": "We the People",
}

incumbency = {
    1912: ("Taft", 1),
    1916: ("Wilson", 1),
    1920: ("Cox", 2),
    1924: ("Coolidge", 1),
    1928: ("Hoover", 2),
    1932: ("Hoover", 1),
    1936: ("Roosevelt", 1),
    1940: ("Roosevelt", 1),
    1944: ("Roosevelt", 1),
    1948: ("Truman", 1),
    1952: ("Stevenson", 2),
    1956: ("Eisenhower", 1),
    1960: ("Nixon", 2),
    1964: ("Johnson", 1),
    1968: ("Humphrey", 2),
    1972: ("Nixon", 1),
    1976: ("Ford", 1),
    1980: ("Carter", 1),
    1984: ("Reagan", 1),
    1988: ("Bush", 2),
    1992: ("Bush", 1),
    1996: ("Clinton", 1),
    2000: ("Gore", 2),
    2004: ("Bush", 1),
    2008: ("McCain", 2),
    2012: ("Obama", 1),
    2016: ("Clinton", 2),
}

data_files = os.listdir(GOV_PATH)
data_files.sort()
for fname in data_files:
    if fname[0] == '.' or fname[0] == "~":
        continue

    year = int(fname[:4])
    if year == 1980:
        print("(Skipping %d president data)" % year)
        continue # missing candidate data from 1980;
                 #  could do manually if need be later
    print("Extracting President voting data from " + str(year))

    fpath = os.path.join(GOV_PATH, fname)
    wb = load_workbook(filename=fpath, data_only=True)

    candidate_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "Candidates":
            candidate_idx = idx
            break
    if candidate_idx == -1:
        print("ERROR! Couldn't find candidates info sheet.")
        sys.exit(1)

    year_candidates = {}
    candidate_data = wb.worksheets[candidate_idx]

    if year in [1984, 1992, 1996, 2000]:
        party_row = 0
        name_row = 2
        if year == 1984:
            party_row = 5
            name_row = 3

        party_cells = candidate_data[party_row+1]
        candidate_cells = candidate_data[name_row+1]

        for col, candidate_cell in enumerate(candidate_cells):
            if col == 0:
                continue
            if len(str(party_cells[col].value).strip()) == 0:
                continue
            if party_cells[col].value == None:
                continue
            candidate = {}
            party_name = party_cells[col].value
            if party_name in equivalents.keys():
                party_name = equivalents[party_name]
            names = candidate_cell.value.split()
            name = names[-1]
            if name == "Jr.":
                name = names[-2]
            elif name == "D.":
                # dealing with "Isabell Masters, pH. D."
                name = names[-3][:-1]
            elif candidate_cell.value == "None of these Candidates":
                name = party_name = "None"
            if party_name not in all_parties.keys():
                all_parties[party_name] = len(all_parties)

            candidate["_idx"] = len(year_candidates)
            candidate["Year"] = year
            candidate["National Party"] = party_name
            candidate["Name"] = name
            candidates.append(candidate)
            year_candidates[name] = candidate

    else:
        short_name_col = -1

        for i, c in enumerate(candidate_data[1]):
            if c.value == "Short Name":
                short_name_col = i
        if short_name_col == -1:
            print("ERROR: Couldn't find short name column.")
            sys.exit(1)

        parties = {}
        for row_idx, curr_row in enumerate(candidate_data.rows):
            if row_idx == 0:
                continue
            if curr_row[0].value == None:
                continue

            party_idx = int(curr_row[0].value)
            party_name = curr_row[3].value
            if party_name == None:
                party_name = "[None]"
            parties[party_idx] = party_name
            if party_name not in all_parties.keys():
                all_parties[party_name] = len(all_parties)

            candidate = {}
            candidate["_idx"] = len(year_candidates)
            candidate["Year"] = year
            candidate["National Party"] = party_name
            candidate["Name"] = curr_row[short_name_col].value
            candidates.append(candidate)
            year_candidates[candidate["Name"]] = candidate

    county_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "County":
            county_idx = idx
            break
    if county_idx == -1:
        print("ERROR! Couldn't find county voting data sheet.")
        sys.exit(1)
    county_data = wb.worksheets[county_idx]

    county_headers = county_data[1]
    fips_column = -1
    for i, c in enumerate(county_headers):
        if c.value == "FIPS":
            fips_column = i
            break
    if fips_column == -1:
        print("ERROR! Couldn't find fips column.")
        sys.exit(1)

    candidate_col = 13 # starting at N

    for curr_row in county_data.iter_rows(min_row=2):
        county = {}
        county_name = curr_row[0].value
        if county_name == None:
            continue
        county_state = curr_row[1].value
        if county_state == "T":
            continue

        county["_year"] = year
        county["_state"] = county_state
        county["_name"] = county_name
        county["_fips"] = curr_row[fips_column].value

        for cname, cdata in year_candidates.items():
            party_name = cdata["National Party"]
            vote_count = curr_row[candidate_col + cdata["_idx"]].value
            if vote_count == None:
                vote_count = 0
            county[party_name] = vote_count

        county_votes.append(county)


candidate_wb = Workbook()
candidate_ws = candidate_wb.active
candidate_ws.title = "President Candidate Data"
candidate_ws.cell(row=1, column=1, value="Year")
candidate_ws.cell(row=1, column=2, value="Incumbency")
candidate_ws.cell(row=1, column=3, value="Name")
candidate_ws.cell(row=1, column=4, value="National Party")

curr_row = 2
for candidate in candidates:
    candidate_ws.cell(row=curr_row, column=1, value=candidate["Year"])
    if incumbency[candidate["Year"]][0] == candidate["Name"]:
        candidate_ws.cell(row=curr_row, column=2, value=incumbency[candidate["Year"]][1])
    else:
        candidate_ws.cell(row=curr_row, column=2, value=0)
    candidate_ws.cell(row=curr_row, column=3, value=candidate["Name"])
    candidate_ws.cell(row=curr_row, column=4, value=candidate["National Party"])
    curr_row += 1

candidate_wb.save(os.path.join(OUTPUT_PATH, "president_candidates.xlsx"))

voting_wb = Workbook()
voting_ws = voting_wb.active
voting_ws.title = "President Voting Data"
voting_ws.cell(row=1, column=1, value="Year")
voting_ws.cell(row=1, column=2, value="State")
voting_ws.cell(row=1, column=3, value="FIPS")
voting_ws.cell(row=1, column=4, value="County Name")

party_col = 5
for pname, pi in all_parties.items():
    voting_ws.cell(row=1, column=party_col + pi, value=pname)

county_row = 2
print("Preparing collective data for output")
for county in tqdm(county_votes):
    voting_ws.cell(row=county_row, column=1, value=county["_year"])
    voting_ws.cell(row=county_row, column=2, value=county["_state"])
    voting_ws.cell(row=county_row, column=3, value=county["_fips"])
    voting_ws.cell(row=county_row, column=4, value=county["_name"])

    party_list = all_parties.copy()
    for col, val in county.items():
        if type(col) == int:
            print(county)
            print(col, val)
        if col[0] == "_":
            continue
        voting_ws.cell(
            row=county_row,
            column=party_col + all_parties[col],
            value=val
        )
        del party_list[col]

    for pname, pi in party_list.items():
        voting_ws.cell(
            row=county_row,
            column = party_col + pi,
            value = 0
        )

    county_row += 1

print("Writing data to output file")
voting_wb.save(os.path.join(OUTPUT_PATH, "president_voting.xlsx"))
