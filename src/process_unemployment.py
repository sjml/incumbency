import os
import sys

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

os.chdir(os.path.dirname(os.path.abspath(__file__)))
BLS_PATH = "../data/raw/bls_lau"
OUTPUT_PATH = "../data/interim/"


def build():
    unemployment = {}
    data_files = os.listdir(BLS_PATH)
    data_files.sort()
    for fname in data_files:
        if fname[0] == '.' or fname[0] == "~":
            continue
        if not fname.endswith(".xlsx"):
            continue

        fpath = os.path.join(BLS_PATH, fname)
        wb = load_workbook(filename=fpath, data_only=True)
        ws = wb.worksheets[0]

        suffix = " Annual Averages"
        year = ws[1][0].value[-len(suffix)-4:-len(suffix)]
        year = int(year)
        print("Extracting unemployment data from %d" % year)

        unemployment[year] = {}

        for row in ws.iter_rows(min_row=7):
            if row[0].value == None:
                break
            state_code = str(row[1].value)
            county_code = str(row[2].value)
            fips = int(state_code + county_code)
            unemployment[year][fips] = (
                row[6].value,
                row[7].value,
                row[8].value,
                row[9].value
            )

    return unemployment

def output(unemployment):
    unemployment_wb = Workbook()
    unemployment_ws = unemployment_wb.active
    unemployment_ws.title = "Unemployment Data"
    unemployment_ws.cell(row=1, column=1,  value="Year")
    unemployment_ws.cell(row=1, column=2,  value="FIPS")
    unemployment_ws.cell(row=1, column=3,  value="Labor Force")
    unemployment_ws.cell(row=1, column=4,  value="Employed")
    unemployment_ws.cell(row=1, column=5,  value="Unemployed")
    unemployment_ws.cell(row=1, column=6,  value="Unemployment Rate")
    unemployment_ws.cell(row=1, column=7,  value="Unemployment Delta_1")
    unemployment_ws.cell(row=1, column=8,  value="Unemployment Delta_2")
    unemployment_ws.cell(row=1, column=9,  value="Unemployment Delta_3")
    unemployment_ws.cell(row=1, column=10, value="Unemployment Delta_4")
    unemployment_ws.cell(row=1, column=11, value="Unemployment Delta_5")
    unemployment_ws.cell(row=1, column=12, value="Unemployment Delta_6")
    unemployment_ws.cell(row=1, column=13, value="Unemployment Delta_7")
    unemployment_ws.cell(row=1, column=14, value="Unemployment Delta_8")
    unemployment_ws.cell(row=1, column=15, value="Unemployment Delta_9")
    unemployment_ws.cell(row=1, column=16, value="Unemployment Delta_10")

    county_row = 2
    for year in sorted(unemployment.keys()):
        for fips_code in sorted(unemployment[year]):
            data = unemployment[year][fips_code]
            unemployment_ws.cell(row=county_row, column=1, value=year)
            unemployment_ws.cell(row=county_row, column=2, value=fips_code)
            unemployment_ws.cell(row=county_row, column=3, value=data[0])
            unemployment_ws.cell(row=county_row, column=4, value=data[1])
            unemployment_ws.cell(row=county_row, column=5, value=data[2])
            unemployment_ws.cell(row=county_row, column=6, value=data[3])

            for offset in range(1,11):
                if (year - offset) in unemployment.keys():
                    if fips_code not in unemployment[(year - offset)]:
                        continue
                    old_data = unemployment[(year - offset)][fips_code]
                    if isinstance(old_data[3], (int, float)) and isinstance(data[3], (int, float)):
                        unemployment_ws.cell(row=county_row, column=6+offset, value=data[3] - old_data[3])

            county_row += 1

    print("Writing data to output file")
    unemployment_wb.save(os.path.join(OUTPUT_PATH, "unemployment_rates.xlsx"))

if __name__ == "__main__":
    data = build()
    output(data)
