# this could easily be accomplished with a simple cut/paste transpose
#   why am I making this more work than it needs to be?!

import os
import sys
import csv
import datetime
import subprocess
from collections import OrderedDict

from openpyxl import Workbook, load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))
GDP_PATH = "../data/raw/world_bank/"
OUTPUT_PATH = "../data/interim/"
GDP_FILE = "API_NY.GDP.MKTP.KD.ZG_DS2_en_excel_v2_422103.xls"
OUTPUT_FILE = "gdp.xlsx"

# requires libreoffice to be installed :-/
if not os.path.exists(OUTPUT_PATH):
    os.makedirs(OUTPUT_PATH)
subprocess.run([
    "soffice",
    "--headless",
    "--convert-to",
    "xlsx",
    os.path.join(GDP_PATH, GDP_FILE),
    "--outdir",
    OUTPUT_PATH
])

fpath = os.path.join(OUTPUT_PATH, GDP_FILE) + "x"
wb = load_workbook(filename=fpath, data_only=True)

data_idx = -1
for idx, sheetname in enumerate(wb.sheetnames):
    if sheetname == "Data":
        data_idx = idx
        break
if data_idx == -1:
    print("ERROR! Couldn't find data sheet.")
    sys.exit(1)

data_sheet = wb.worksheets[data_idx]

output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "GDP Growth"
output_ws.cell(row=1, column=1, value="Year")
output_ws.cell(row=1, column=2, value="GDP_Growth")

# data starts at column 6 (F), labels in row 4; US in row 254
out_row = 2
for cell in data_sheet[254]:
    if cell.column < 6:
        continue
    year = data_sheet.cell(row=4, column=cell.column).value
    gdp_growth = cell.value
    output_ws.cell(row=out_row, column=1, value=year)
    output_ws.cell(row=out_row, column=2, value=gdp_growth)
    out_row += 1

output_wb.save(os.path.join(OUTPUT_PATH, OUTPUT_FILE))
