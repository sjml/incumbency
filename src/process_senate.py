import os
import sys
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

os.chdir(os.path.dirname(os.path.abspath(__file__)))
SENATE_PATH = "../data/raw/david_leip/Senate Election Data (xls)"
OUTPUT_PATH = "../data/interim/"

candidates = []
county_votes = []
all_parties = OrderedDict()

data_files = os.listdir(SENATE_PATH)
data_files.sort()
for fname in data_files:
    if fname[0] == '.' or fname[0] == "~":
        continue
    year = int(fname[len("Sen_Election_Data_"):len("Sen_Election_Data_")+4])
    print("Extracting Senate voting data from " + str(year))

    fpath = os.path.join(SENATE_PATH, fname)
    wb = load_workbook(filename=fpath, data_only=True)

    candidate_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "Candidates":
            candidate_idx = idx
            break
    if candidate_idx == -1:
        print("ERROR! Couldn't find candidate info sheet.")
        sys.exit(1)

    blanks_allowed = 20
    candidate_data = wb.worksheets[candidate_idx]

    ok = False
    short_name_col = 1
    blanks_allowed = 20
    while blanks_allowed > 0:
        val = candidate_data.cell(row=1, column=short_name_col).value
        if val == "Short Name":
            ok = True
            break
        short_name_col += 1

    if not ok:
        print("ERROR: Couldn't find short name column.")
        sys.exit(1)

    party_row = 2
    parties = {}
    while True:
        party_idx = candidate_data.cell(row=party_row, column=1).value
        if party_idx == None:
            break
        party_idx = int(party_idx)
        party_name = candidate_data.cell(row=party_row, column=short_name_col).value
        if party_name == None:
            party_name = "[None]"
        # The data set used different short names for the Peace and Freedom party across years;
        # normalizing it
        if party_name in ["Peace&Free", "Peace & Free"]:
            party_name = "Peace and Free"
        parties[party_idx] = party_name
        if party_name not in all_parties.keys():
            all_parties[party_name] = len(all_parties)
        party_row += 1

    ok = False
    candidate_row_start = party_row

    while blanks_allowed > 0:
        val = candidate_data.cell(row=candidate_row_start, column=1).value
        if val == None:
            candidate_row_start += 1
            blanks_allowed -= 1
            continue
        val = str(val)
        if len(val) == 0:
            candidate_row_start += 1
            blanks_allowed -= 1
            continue
        if val.strip() == "#":
            ok = True
            break
        candidate_row_start += 1

    if not ok:
        print("ERROR: Couldn't find candidate region. " + str(candidate_row_start))
        sys.exit(1)

    incumbency_col = short_name_col + 2

    curr_row = candidate_row_start + 1
    while True:
        candidate = {}
        natl_party_idx = candidate_data.cell(row=curr_row, column=1).value
        if natl_party_idx == None:
            break
        candidate["Year"] = year
        candidate["National Party"] = parties[int(natl_party_idx)]
        candidate["Name"] = candidate_data.cell(row=curr_row, column=2).value
        candidate["Ballot Party"] = candidate_data.cell(row=curr_row, column=3).value
        candidate["State"] = candidate_data.cell(row=curr_row, column=4).value
        candidate["Incumbency"] = candidate_data.cell(row=curr_row, column=incumbency_col).value
        candidates.append(candidate)
        curr_row += 1


    county_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "County":
            county_idx = idx
            break
    if county_idx == -1:
        print("ERROR! Couldn't find county voting data sheet.")
        sys.exit(1)
    county_data = wb.worksheets[county_idx]

    ok = False
    fips_column = 1
    while fips_column < county_data.max_column:
        if county_data.cell(row=1, column=fips_column).value == "FIPS":
            ok = True
            break
        fips_column += 1
    if ok == False:
        print("ERROR! Couldn't find FIPS column.")
        sys.exit(1)

    party_col = 14 # starting at N
    if county_data.cell(row=1, column=party_col).value != "Democratic":
        print("ERROR! Parties don't start at expected column.")
        sys.exit(1)
    for pi in range(len(parties)):
        pbase = parties[pi+1]
        pcheck = county_data.cell(row=1, column=party_col + pi).value
        if pcheck == 0:
            pcheck = "[None]"
        if pcheck in ["Peace&Free", "Peace & Free"]:
            pcheck = "Peace and Free"
        if pbase != pcheck:
            print("ERROR! Party", pbase, "doesn't match", pcheck, "at index", pi)
            sys.exit(1)

    for row in county_data.iter_rows(min_row=2):
        county = {}
        county_name = row[0].value
        if county_name == None:
            continue
        county_state = row[1].value
        if county_state == "T":
            continue

        county["_year"] = year
        county["_state"] = county_state
        county["_name"] = county_name
        county["_fips"] = row[fips_column-1].value

        for pi in range(len(parties)):
            vote_count = row[party_col-1 + pi].value
            if vote_count == None or len(str(vote_count).strip()) == 0:
                vote_count = 0
            county[parties[pi+1]] = vote_count

        county_votes.append(county)


candidate_wb = Workbook()
candidate_ws = candidate_wb.active
candidate_ws.title = "Senate Candidate Data"
candidate_ws.cell(row=1, column=1, value="Year")
candidate_ws.cell(row=1, column=2, value="State")
candidate_ws.cell(row=1, column=3, value="Incumbency")
candidate_ws.cell(row=1, column=4, value="Name")
candidate_ws.cell(row=1, column=5, value="National Party")
candidate_ws.cell(row=1, column=6, value="Ballot Party")

curr_row = 2
for candidate in candidates:
    candidate_ws.cell(row=curr_row, column=1, value=candidate["Year"])
    candidate_ws.cell(row=curr_row, column=2, value=candidate["State"])
    candidate_ws.cell(row=curr_row, column=3, value=candidate["Incumbency"])
    candidate_ws.cell(row=curr_row, column=4, value=candidate["Name"])
    candidate_ws.cell(row=curr_row, column=5, value=candidate["National Party"])
    candidate_ws.cell(row=curr_row, column=6, value=candidate["Ballot Party"])
    curr_row += 1

candidate_wb.save(os.path.join(OUTPUT_PATH, "senate_candidates.xlsx"))


voting_wb = Workbook()
voting_ws = voting_wb.active
voting_ws.title = "Senate Voting Data"
voting_ws.cell(row=1, column=1, value="Year")
voting_ws.cell(row=1, column=2, value="State")
voting_ws.cell(row=1, column=3, value="FIPS")
voting_ws.cell(row=1, column=4, value="County Name")

party_col = 5
for pname, pi in all_parties.items():
    voting_ws.cell(row=1, column=party_col + pi, value=pname)

county_row = 2
print("Preparing collective data for output")
for county in tqdm(county_votes):
    voting_ws.cell(row=county_row, column=1, value=county["_year"])
    voting_ws.cell(row=county_row, column=2, value=county["_state"])
    voting_ws.cell(row=county_row, column=3, value=county["_fips"])
    voting_ws.cell(row=county_row, column=4, value=county["_name"])

    party_list = all_parties.copy()
    for col, val in county.items():
        if col[0] == "_":
            continue
        voting_ws.cell(
            row=county_row,
            column=party_col + all_parties[col],
            value=val
        )
        del party_list[col]

    for pname, pi in party_list.items():
        voting_ws.cell(
            row=county_row,
            column = party_col + pi,
            value = 0
        )

    county_row += 1

print("Writing data to output file")
voting_wb.save(os.path.join(OUTPUT_PATH, "senate_voting.xlsx"))
