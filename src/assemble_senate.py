import os
import sys

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

import state_lookup

os.chdir(os.path.dirname(os.path.abspath(__file__)))
MARKET_HISTORY_DATA_FILE = "../data/interim/market_history.xlsx"
GDP_GROWTH_DATA_FILE = "../data/interim/gdp.xlsx"
UNEMPLOYMENT_DATA_FILE = "../data/interim/unemployment_rates.xlsx"
SENATE_CANDIDATES_FILE = "../data/interim/senate_candidates.xlsx"
SENATE_VOTING_FILE = "../data/interim/senate_voting.xlsx"
SENATE_OUTPUT_FILE = "../data/processed/senate_data.xlsx"

market_history = {}
market_history_wb = load_workbook(filename=MARKET_HISTORY_DATA_FILE, data_only=True)
market_history_ws = market_history_wb.active
market_history_headers = [x.value for x in market_history_ws[1][1:]]
for row in market_history_ws.iter_rows(min_row=2):
    year = int(row[0].value)
    market_history[year] = [x.value for x in row[1:]]

gdp_growth = {}
gdp_growth_wb = load_workbook(filename=GDP_GROWTH_DATA_FILE, data_only=True)
gdp_growth_ws = gdp_growth_wb.active
for row in gdp_growth_ws.iter_rows(min_row=2):
    year = int(row[0].value)
    gdp_growth[year] = row[1].value

# year -> fips -> (unemployment_data)
unemployment = {}
unemployment_wb = load_workbook(filename=UNEMPLOYMENT_DATA_FILE, data_only=True)
unemployment_ws = unemployment_wb.active

unemployment_header_cells = unemployment_ws[1]
unemployment_headers = {}
for i, c in enumerate(unemployment_header_cells):
    unemployment_headers[c.value] = i

for row in unemployment_ws.iter_rows(min_row=2):
    year = row[unemployment_headers["Year"]].value
    fips = row[unemployment_headers["FIPS"]].value
    labor_force = row[unemployment_headers["Labor Force"]].value
    employed = row[unemployment_headers["Employed"]].value
    unemployed = row[unemployment_headers["Unemployed"]].value
    unemployment_rate = row[unemployment_headers["Unemployment Rate"]].value
    unemployment_deltas = {}
    for offset in range(1, 11):
        unemployment_deltas[offset] = row[unemployment_headers["Unemployment Delta_%d" % offset]].value

    if year not in unemployment.keys():
        unemployment[year] = {}
    unemployment[year][fips] = (
        labor_force, employed, unemployed, unemployment_rate, unemployment_deltas
    )
unemployment_wb.close()

# year -> state -> (incumbent name, incumbent candidate's party, incumbency_type)
senate_elections = {}
senate_candidates_wb = load_workbook(filename=SENATE_CANDIDATES_FILE, data_only=True)
senate_candidates_ws = senate_candidates_wb.active

senate_header_cells = senate_candidates_ws[1]
senate_headers = {}
for i, c in enumerate(senate_header_cells):
    senate_headers[c.value] = i

for row in senate_candidates_ws.iter_rows(min_row=2):
    inc = row[senate_headers["Incumbency"]].value
    if inc is not None and inc <= 0:
        continue
    year = row[senate_headers["Year"]].value
    state = row[senate_headers["State"]].value
    state = state_lookup.abbreviation_from_name(state)

    if year not in senate_elections.keys():
        senate_elections[year] = {}
    if state not in senate_elections[year].keys():
        senate_elections[year][state] = (
            row[senate_headers["Name"]].value,
            row[senate_headers["National Party"]].value,
            row[senate_headers["Incumbency"]].value
        )

senate_candidates_wb.close()


senate_voting_wb = load_workbook(filename=SENATE_VOTING_FILE, data_only=True)
senate_voting_ws = senate_voting_wb.active

senate_voting_header_cells = senate_voting_ws[1]
senate_voting_headers = {}
for i, c in enumerate(senate_voting_header_cells):
    senate_voting_headers[c.value] = i


senate_wb = Workbook()
senate_ws = senate_wb.active
senate_ws.title = "Senate Inc. vs. Unemployment"
senate_ws.cell(row=1, column=1, value="Year")
senate_ws.cell(row=1, column=2, value="FIPS")
senate_ws.cell(row=1, column=3, value="State")
senate_ws.cell(row=1, column=4, value="County")
senate_ws.cell(row=1, column=5, value="Incumbent Name")
senate_ws.cell(row=1, column=6, value="Incumbent Party")
senate_ws.cell(row=1, column=7, value="Incumbency Code")
senate_ws.cell(row=1, column=8, value="Votes For Incumbent")
senate_ws.cell(row=1, column=9, value="Votes Against Incumbent")
senate_ws.cell(row=1, column=10, value="GDP Growth")
senate_ws.cell(row=1, column=11, value="Labor Force")
senate_ws.cell(row=1, column=12, value="Employed")
senate_ws.cell(row=1, column=13, value="Unemployed")
senate_ws.cell(row=1, column=14, value="Unemployment Rate")
curr_col = 15
for offset in range(1, 11):
    senate_ws.cell(row=1, column=curr_col, value="Unemployment Delta_%d" % offset)
    curr_col += 1
for market_index_info in market_history_headers:
    senate_ws.cell(row=1, column=curr_col, value=market_index_info)
    curr_col += 1

discard_count = 0
total_count = 0
out_row = 2
for row in senate_voting_ws.iter_rows(min_row=2):
    total_count += 1
    year = row[senate_voting_headers["Year"]].value
    state = row[senate_voting_headers["State"]].value
    fips = row[senate_voting_headers["FIPS"]].value

    if state == None:
        discard_count += 1
        continue
    try:
        unemployment_data = unemployment[year][fips]
    except KeyError:
        discard_count += 1
        continue

    senate_ws.cell(row=out_row, column=1, value=year)
    senate_ws.cell(row=out_row, column=2, value=fips)
    senate_ws.cell(row=out_row, column=3, value=state)
    senate_ws.cell(row=out_row, column=4, value=row[senate_voting_headers["County Name"]].value)

    incumbent_data = senate_elections[year][state]
    senate_ws.cell(row=out_row, column=5, value=incumbent_data[0])
    senate_ws.cell(row=out_row, column=6, value=incumbent_data[1])
    senate_ws.cell(row=out_row, column=7, value=incumbent_data[2])

    inc_votes = 0
    opp_votes = 0
    try:
        for i in range(4, len(row)):
            if senate_voting_header_cells[i].value == incumbent_data[1]:
                inc_votes += int(row[i].value)
            else:
                opp_votes += int(row[i].value)
    except Exception as exc:
        print("ERROR!", total_count)
        raise exc

    senate_ws.cell(row=out_row, column=8, value=inc_votes)
    senate_ws.cell(row=out_row, column=9, value=opp_votes)

    senate_ws.cell(row=out_row, column=10, value=gdp_growth[year])

    senate_ws.cell(row=out_row, column=11, value=unemployment_data[0])
    senate_ws.cell(row=out_row, column=12, value=unemployment_data[1])
    senate_ws.cell(row=out_row, column=13, value=unemployment_data[2])
    senate_ws.cell(row=out_row, column=14, value=unemployment_data[3])

    curr_col = 15
    for offset in range(1, 11):
        senate_ws.cell(row=out_row, column=curr_col, value=unemployment_data[4][offset])
        curr_col += 1
    market_data = market_history[year]
    for datum in market_data:
        senate_ws.cell(row=out_row, column=curr_col, value=datum)
        curr_col += 1

    out_row += 1

if discard_count > 0:
    print("Discarded %d data points out of %d for lacking matching data" % (discard_count, total_count))
else:
    print("%d data points output" % (total_count))

senate_wb.save(SENATE_OUTPUT_FILE)
