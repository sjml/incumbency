import os
import sys
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

os.chdir(os.path.dirname(os.path.abspath(__file__)))
GOV_PATH = "../data/interim/governor_data/"
OUTPUT_PATH = "../data/processed/"

candidates = []
county_votes = []
all_parties = OrderedDict()

# here are separate codes for parties that should
#  be the same (just different abbreviations across years)
equivalents = {
    0: "[None]",
    "Amer. Ind.": "American Ind.",
    "Democratic": "Democrat",
    "For The People": "For the People",
    "Independ.": "Independent",
    "Lberty Union": "Liberty Union",
    "Nat. Law": "Natural Law",
    "P & F": "Peace & Free",
    "Write-in": "Write-ins",
}

data_files = os.listdir(GOV_PATH)
data_files.sort()
for fname in data_files:
    if fname[0] == '.' or fname[0] == "~":
        continue

    year = int(fname[:4])
    if year == 2003:
        print("(Skipping 2003 governor data)")
        continue # 2003 has significant coding errors
                 # on the candidates page, so unable to
                 # extract party affiliations or
                 # incumbency. Skipping.
    print("Extracting Governor voting data from " + str(year))

    fpath = os.path.join(GOV_PATH, fname)
    wb = load_workbook(filename=fpath, data_only=True)

    candidate_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "Candidates":
            candidate_idx = idx
            break
    if candidate_idx == -1:
        print("ERROR! Couldn't find candidates info sheet.")
        sys.exit(1)

    candidate_data = wb.worksheets[candidate_idx]

    short_name_col = -1
    governor_col = -1

    for i, c in enumerate(candidate_data[1]):
        if c.value == "Short Name":
            short_name_col = i
        if c.value == "Governor":
            governor_col = i
    if short_name_col == -1:
        print("ERROR: Couldn't find short name column.")
        sys.exit(1)
    if governor_col == -1:
        print("ERROR: Couldn't find Governor column.")
        sys.exit(1)

    incumbent_col = short_name_col + 2

    parties = {}
    in_party_section = True
    for row_idx, curr_row in enumerate(candidate_data.rows):
        if row_idx == 0: continue

        party_idx = curr_row[0].value
        if party_idx == None:
            in_party_section = False
            continue
        party_idx = int(party_idx)
        if in_party_section:
            party_name = curr_row[short_name_col].value
            if party_name == None:
                party_name = "[None]"
            if party_name in equivalents.keys():
                party_name = equivalents[party_name]
            parties[party_idx] = party_name
            if party_name not in all_parties.keys():
                all_parties[party_name] = len(all_parties)

        else:
            no_incumbency_years = [1996, 1998, 1999, 2000, 2001, 2002]
            candidate = {}
            natl_party_idx = curr_row[0].value
            if natl_party_idx == None:
                break
            candidate["Year"] = year
            candidate["National Party"] = parties[int(natl_party_idx)]
            candidate["Name"] = curr_row[short_name_col].value
            candidate["Ballot Party"] = curr_row[3].value
            candidate["State"] = curr_row[4].value
            if year in no_incumbency_years:
                candidate["Incumbency"] = -1
            else:
                candidate["Incumbency"] = curr_row[incumbent_col].value
            candidates.append(candidate)

    county_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "County":
            county_idx = idx
            break
    if county_idx == -1:
        print("ERROR! Couldn't find county voting data sheet.")
        sys.exit(1)
    county_data = wb.worksheets[county_idx]

    county_headers = county_data[1]
    fips_column = -1
    for i, c in enumerate(county_headers):
        if c.value == "FIPS":
            fips_column = i
            break
    if fips_column == -1:
        print("ERROR! Couldn't find fips column.")
        sys.exit(1)

    party_col = 13 # starting at N
    first_party = county_headers[party_col].value
    if first_party in equivalents.keys():
        first_party = equivalents[first_party]
    if first_party != "Democrat":
        print("ERROR! Parties don't start at expected column.")
        sys.exit(1)
    for pi in range(len(parties)):
        pbase = parties[pi+1]
        pcheck = county_headers[party_col + pi].value

        if pcheck == 0:
            pcheck = "[None]"
        if pcheck in equivalents.keys():
            pcheck = equivalents[pcheck]
        if pbase != pcheck:
            print("ERROR! Party", pbase, "doesn't match", pcheck, "at index", pi)
            sys.exit(1)

    for curr_row in county_data.iter_rows(min_row=2):
        county = {}
        county_name = curr_row[0].value
        if county_name == None:
            continue
        county_state = curr_row[1].value
        if county_state == "T":
            continue

        county["_year"] = year
        county["_state"] = county_state
        county["_name"] = county_name
        county["_fips"] = curr_row[fips_column].value

        for pi in range(len(parties)):
            if parties[pi+1] == 0:
                continue # empty party name in candidates page
                         # still counting to keep the index
            vote_count = curr_row[party_col + pi].value
            if vote_count == None or len(str(vote_count).strip()) == 0:
                vote_count = 0
            county[parties[pi+1]] = vote_count

        county_votes.append(county)


candidate_wb = Workbook()
candidate_ws = candidate_wb.active
candidate_ws.title = "Governor Candidate Data"
candidate_ws.cell(row=1, column=1, value="Year")
candidate_ws.cell(row=1, column=2, value="State")
candidate_ws.cell(row=1, column=3, value="Incumbency")
candidate_ws.cell(row=1, column=4, value="Name")
candidate_ws.cell(row=1, column=5, value="National Party")
candidate_ws.cell(row=1, column=6, value="Ballot Party")

curr_row = 2
for candidate in candidates:
    candidate_ws.cell(row=curr_row, column=1, value=candidate["Year"])
    candidate_ws.cell(row=curr_row, column=2, value=candidate["State"])
    candidate_ws.cell(row=curr_row, column=3, value=candidate["Incumbency"])
    candidate_ws.cell(row=curr_row, column=4, value=candidate["Name"])
    candidate_ws.cell(row=curr_row, column=5, value=candidate["National Party"])
    candidate_ws.cell(row=curr_row, column=6, value=candidate["Ballot Party"])
    curr_row += 1

candidate_wb.save(os.path.join(OUTPUT_PATH, "governor_candidates.xlsx"))


voting_wb = Workbook()
voting_ws = voting_wb.active
voting_ws.title = "Senate Voting Data"
voting_ws.cell(row=1, column=1, value="Year")
voting_ws.cell(row=1, column=2, value="State")
voting_ws.cell(row=1, column=3, value="FIPS")
voting_ws.cell(row=1, column=4, value="County Name")

party_col = 5
for pname, pi in all_parties.items():
    voting_ws.cell(row=1, column=party_col + pi, value=pname)

county_row = 2
print("Preparing collective data for output")
for county in tqdm(county_votes):
    voting_ws.cell(row=county_row, column=1, value=county["_year"])
    voting_ws.cell(row=county_row, column=2, value=county["_state"])
    voting_ws.cell(row=county_row, column=3, value=county["_fips"])
    voting_ws.cell(row=county_row, column=4, value=county["_name"])

    party_list = all_parties.copy()
    for col, val in county.items():
        if type(col) == int:
            print(county)
            print(col, val)
        if col[0] == "_":
            continue
        voting_ws.cell(
            row=county_row,
            column=party_col + all_parties[col],
            value=val
        )
        del party_list[col]

    for pname, pi in party_list.items():
        voting_ws.cell(
            row=county_row,
            column = party_col + pi,
            value = 0
        )

    county_row += 1

print("Writing data to output file")
voting_wb.save(os.path.join(OUTPUT_PATH, "governor_voting.xlsx"))
