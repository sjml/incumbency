import os
import sys

from openpyxl import Workbook, load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))
SENATE_PATH = "../data/raw/david_leip/Senate Election Data (xls)"
OUTPUT_PATH = "../data/processed/"

candidates = []

data_files = os.listdir(SENATE_PATH)
data_files.sort()
for fname in data_files:
    if fname[0] == '.' or fname[0] == "~":
        continue
    year = int(fname[len("Sen_Election_Data_"):len("Sen_Election_Data_")+4])
    print("Analyzing data from " + str(year))

    fpath = os.path.join(SENATE_PATH, fname)
    wb = load_workbook(filename=fpath, data_only=True)

    candidate_idx = -1
    for idx, sheetname in enumerate(wb.sheetnames):
        if sheetname == "Candidates":
            candidate_idx = idx
            break
    if candidate_idx == -1:
        print("ERROR! Couldn't find candidate info sheet.")
        sys.exit(1)

    blanks_allowed = 20
    candidate_data = wb.worksheets[candidate_idx]

    party_row = 2
    parties = {}
    while True:
        party_idx = candidate_data.cell(row=party_row, column=1).value
        if party_idx == None:
            break
        party_idx = int(party_idx)
        parties[party_idx] = candidate_data.cell(row=party_row, column=3).value
        party_row += 1

    ok = False
    candidate_row_start = party_row

    while blanks_allowed > 0:
        val = candidate_data.cell(row=candidate_row_start, column=1).value
        if val == None:
            candidate_row_start += 1
            blanks_allowed -= 1
            continue
        val = str(val)
        if len(val) == 0:
            candidate_row_start += 1
            blanks_allowed -= 1
            continue
        if val.strip() == "#":
            ok = True
            break
        candidate_row_start += 1

    if not ok:
        print("ERROR: Couldn't find candidate region. " + str(candidate_row_start))
        sys.exit(1)

    ok = False
    short_name_col = 1
    blanks_allowed = 20
    while blanks_allowed > 0:
        val = candidate_data.cell(row=candidate_row_start, column=short_name_col).value
        if val == "Short Name":
            ok = True
            break
        short_name_col += 1

    if not ok:
        print("ERROR: Couldn't find incumbency column.")
        sys.exit(1)

    incumbency_col = short_name_col + 2

    curr_row = candidate_row_start + 1
    while True:
        candidate = {}
        natl_party_idx = candidate_data.cell(row=curr_row, column=1).value
        if natl_party_idx == None:
            break
        candidate["Year"] = year
        candidate["National Party"] = parties[int(natl_party_idx)]
        candidate["Name"] = candidate_data.cell(row=curr_row, column=2).value
        candidate["Ballot Party"] = candidate_data.cell(row=curr_row, column=3).value
        candidate["State"] = candidate_data.cell(row=curr_row, column=4).value
        candidate["Incumbency"] = candidate_data.cell(row=curr_row, column=incumbency_col).value
        candidates.append(candidate)
        curr_row += 1



outwb = Workbook()
outws = outwb.active
outws.title = "Senate Candidate Data"
outws.cell(row=1, column=1, value="Year")
outws.cell(row=1, column=2, value="State")
outws.cell(row=1, column=3, value="Incumbency")
outws.cell(row=1, column=4, value="Name")
outws.cell(row=1, column=5, value="National Party")
outws.cell(row=1, column=6, value="Ballot Party")

curr_row = 2
for candidate in candidates:
    outws.cell(row=curr_row, column=1, value=candidate["Year"])
    outws.cell(row=curr_row, column=2, value=candidate["State"])
    outws.cell(row=curr_row, column=3, value=candidate["Incumbency"])
    outws.cell(row=curr_row, column=4, value=candidate["Name"])
    outws.cell(row=curr_row, column=5, value=candidate["National Party"])
    outws.cell(row=curr_row, column=6, value=candidate["Ballot Party"])
    curr_row += 1

outwb.save(os.path.join(OUTPUT_PATH, "senate_candidates.xlsx"))
